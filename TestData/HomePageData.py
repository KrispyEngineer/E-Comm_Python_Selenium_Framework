import openpyxl



class HomePageData:

    #test_HomePage_data = [{'firstname':'Piyush', 'lastname':'Patil', 'password': '123456', 'gender':'Male'}, {'firstname':'xyz', 'lastname':'pqr', 'password': '1234', 'gender':'Female'} ]


    @staticmethod
    def getExcelData(test_case_name):
        book = openpyxl.load_workbook("C:\\Users\\piyus\\Desktop\\lp3thw\\Python Testing\\PythonSelFramework\\Resources\\input_data.xlsx")
        sheet = book.active
        Dict = {}
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [Dict]
